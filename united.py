import pandas as pd
from email.mime.text import MIMEText
import smtplib
import sys
import time
import imaplib
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
import pdftotext
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import html2text
import subprocess

from decode_error import check_subject, read_from_delete
from make_log import log_exceptions

try:
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])

    config = pdfkit.configuration(wkhtmltopdf='/usr/bin/wkhtmltopdf')
    if path.exists(r'united/united' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'united/united' + str(sys.argv[6]) + '.xlsx')
    import openpyxl

    po = []
    fg = []
    repeat = []
    wbkName = 'united/united' + str(sys.argv[6]) + '.xlsx'


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        # mail.login(user = 'Mediclaim@therisingmedicare.com', password = 'cef@2018')

        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None, '(SUBJECT "Settlement voucher for LOC" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None, '(SUBJECT "Settlement voucher for LOC" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        #
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id, "(RFC822)")
            ##################################################ak
            try:
                raw_email = data[0][1].decode('utf-8')
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            # if path.exists(r'/home/shivam/Desktop/vnu_scripts/Paramount/email.html'):
            # os.remove(r'email.html')
            # Body details
            if email_message['Subject'] not in fg:
                for part in email_message.walk():
                    # print(part.get_content_type())
                    if part.get_content_type() == "text/html":
                        # print('hi')
                        body = part.get_payload(decode=True)
                        file_name = "united/email.html"
                        output_file = open(file_name, 'w')
                        output_file.write("Body: %s" % (body.decode('utf-8')))
                        output_file.close()
                        pdfkit.from_file('united/email.html',
                                         'united/attachments_pdf_' + str(sys.argv[6]) + '/' + str(i) + '.pdf',configuration=config)
                    else:
                        continue
            else:
                repeat.append(email_message['Subject'])
            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/united'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_pdf_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_pdf_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/united/attachments_pdf_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    wbk.create_sheet('2')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    s3 = wbk.worksheets[2]

    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    for t in range(0, len(onlyfiles)):
        try:
            s3.cell(row=1, column=1).value = 'Sr. No.'
            s3.cell(row=1, column=2).value = 'Claim No/URL No'
            s3.cell(row=1, column=3).value = 'Amount'
            s3.cell(row=1, column=4).value = 'remark'
            with open(mypath + onlyfiles[t], "rb") as f:
                pdf = pdftotext.PDF(f)

            with open('united/output.txt', 'w') as f:
                f.write(" ".join(pdf))
            with open('united/output.txt', 'r') as myfile:
                f = myfile.read()
            # print(f)

            sh2 = ['Policy Number', 'Employee No.', 'Patient Name', 'Insurance Company', 'EFT Transfer Code', 'DOA', 'DOD',
                   'Loc No', 'Diagnosis', 'LOC/AL Amount', 'Claim Amount', 'ChequeDate', 'primary benifiary']
            for i in range(0, len(sh2)):
                s1.cell(row=1, column=i + 3).value = sh2[i]
            sh1 = ['UHC Approved Hospital Amt', 'UHC Approved Employee Amt', 'Insurer Approved Hospital Amt',
                   'Insurer Approved Employee Amt', 'Payable Amount Hospital Amt', 'Payable Amount Employee Amt', 'TDS']

            for i in range(0, len(sh1)):
                s2.cell(row=1, column=i + 3).value = sh1[i]
            hg = []
            w = f.find('Policy No.') + 10
            g = f[w:]
            u = g.find('\n') + w
            hg.append(f[w:u])

            w1 = f.find('Employee No.') + 12
            g = f[w1:]
            u1 = g.find('\n') + w1
            hg.append(f[w1:u1])

            w2 = f.find('Patient') + 7
            g = f[w2:]
            u2 = g.find('\n') + w2
            hg.append(f[w2:u2])

            w3 = f.find('policyholder of the') + 19
            g = f[w3:]
            u3 = g.find(',') + w3
            hg.append(f[w3:u3])

            w4 = f.find('EFT Transfer Code') + 17
            g = f[w4:]
            u4 = g.find('\n') + w4
            hg.append(f[w4:u4])

            w5 = f.find('DOA') + 4
            g = f[w5:]
            u5 = g.find('-') + w5
            hg.append(f[w5:u5])

            w6 = f.find('DOD') + 4
            g = f[w6:]
            u6 = g.find('\n') + w6
            hg.append(f[w6:u6])

            w7 = f.find('Loc No') + 7
            g = f[w7:]
            u7 = g.find('\n') + w7
            hg.append(f[w7:u7])

            w8 = f.find('Disease') + 7
            g = f[w8:]
            u8 = g.find('\n') + w8
            hg.append(f[w8:u8])

            w8 = f.find('LOC/AL Amount') + 14
            g = f[w8:]
            u8 = g.find('\n') + w8
            hg.append(f[w8:u8])

            w8 = f.find('Claim Amount') + 13
            g = f[w8:]
            u8 = g.find('\n') + w8
            hg.append(f[w8:u8])

            w8 = f.find('/Date') + 6
            g = f[w8:]
            u8 = g.find('-') + w8
            hg.append(f[w8:u8])

            w2 = f.find('/Employee') + 10
            g = f[w2:]
            u2 = g.find('\n') + w2
            hg.append(f[w2:u2])

            w9 = f.find('Claim No/URL No') + 15
            g = f[w9:]
            u9 = g.find('Loc') + w9
            ccn = (f[w9:u9])
            r5 = ccn.find('/') + 1
            ccn = ccn[r5:]
            ccn = ccn.replace('  ', '')
            hg = [sub.replace('  ', '') for sub in hg]
            hg = [sub.replace(':', '') for sub in hg]
            hg = [sub.replace('\n', ' ') for sub in hg]
            for i in range(0, len(hg)):
                s1.cell(row=t + 2, column=i + 3).value = hg[i]
            gh = []
            w = f.find('UHC Approved')
            # print(w)
            g = f[w:]
            w1 = g.find(':') + w + 1
            u = g.find('.') + w + 3
            gh.append(f[w1:u])
            w1 = g.find('Employee Amt') + w + 14
            u1 = g.find('\n') + w
            gh.append(f[w1:u1])

            w2 = f.find('Insurer Approved')
            g = f[w2:]
            w3 = g.find(':') + w2 + 1
            u2 = g.find('.') + w2 + 3
            gh.append(f[w3:u2])
            # print(w3,u2)

            w3 = g.find('Employee Amt') + w2 + 14
            u3 = g.find('\n') + w2
            gh.append(f[w3:u3])

            w4 = f.find('Payable Amount')
            g = f[w4:]
            w5 = g.find(':') + w4 + 1
            u4 = g.find('.') + w4 + 3
            gh.append(f[w5:u4])

            w5 = g.find('Employee Amt') + w4 + 14
            u5 = g.find('\n') + w4
            gh.append(f[w5:u5])
            # print(gh)
            if f.find('TDS:') != -1:
                w5 = f.find('TDS:') + 4
                g = f[w5:]
                u5 = g.find(']') + w5
                gh.append(f[w5:u5])
            else:
                gh.append(' ')
            # print(gh)
            for i in range(0, len(gh)):
                gh[i] = gh[i].replace('', "")
                s2.cell(row=t + 2, column=i + 3).value = gh[i].replace('', "")
            for wd in wbk.worksheets[:2]:
                wd.cell(row=1, column=1).value = 'Sr. No.'
                wd.cell(row=1, column=2).value = 'Claim No/URL No'
                wd.cell(row=t + 2, column=1).value = t + 1
                wd.cell(row=t + 2, column=2).value = ccn

            w5 = f.find('Cheque No/Date')
            g = f[w5:]
            w6 = g.find('\n') + w5
            u5 = g.find('Encashment') + w5
            we = f[w6:u5]
            we = we.replace('Deductions/Remarks', '')
            we = we.replace('  ', '')
            we = we.replace(' +', '+')
            we = we.replace('\n', '$$')
            we = we.replace('$$$$', '$$')
            # print(we)
            op = we.split('$$+')
            op = [i.replace('$$', '') for i in op]
            # print(op)
            ro = []
            yo = []
            for j in op:
                i = j.find('/-')
                if i != -1:
                    ro.append(j[:i])
                    yo.append(j[i + 2:])
                else:
                    ro.append('0')
                    yo.append('-')

            for i in range(0, len(ro)):
                row_num = s3.max_row + 1
                if yo[i] == 'Advance()':
                    continue
                else:
                    s3.cell(row=row_num, column=2).value = ccn
                    s3.cell(row=row_num, column=3).value = ro[i]
                    s3.cell(row=row_num, column=4).value = yo[i]
            # print(ro,yo)
            ccn = ccn.replace(' ', '')
            temp = ccn.replace('/', '')
            os.rename(os.getcwd() + '/united/attachments_pdf_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/united/attachments_pdf_' + str(sys.argv[6]) + '/' + temp + '.pdf')
        except Exception as e:
            log_exceptions()
            s1.cell(row=t + 2, column=1).value = 'error'
            os.rename(os.getcwd() + '/united/attachments_pdf_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/united/attachments_pdf_' + str(sys.argv[6]) + '/' + temp + '.pdf')

    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'united'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
	log_exceptions()