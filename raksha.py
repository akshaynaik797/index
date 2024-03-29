import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import sys
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
import pdftotext
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import html2text
import subprocess

from decode_error import check_subject, read_from_delete
from make_log import log_exceptions

try:

    fg = []
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])
    repeat = []
    wq = 0


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(SUBJECT "Claim Settlement Letter From Raksha Health Insurance TPA Pvt.Ltd." since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None,
        #                          '(SUBJECT "Claim Settlement Letter From Raksha Health Insurance TPA Pvt.Ltd." since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id,
                                      "(RFC822)")  # fetch the email body (RFC822)             for the given ID

            ##################################################ak
            try:
                raw_email = data[0][1].decode('utf-8')
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend

            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():
                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    if part.get('Content-Disposition') is None:
                        # print part.as_string()
                        continue
                    fileName = part.get_filename()
                    detach_dir = (os.getcwd() + '/raksha/attachments_' + str(sys.argv[6]))
                    if bool(fileName):
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName + '.pdf')
                            # print (fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
            else:
                repeat.append(email_message['Subject'])

            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/raksha'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/raksha/attachments_' + str(sys.argv[6]) + '/'

    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    import sys

    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        # print(filename)
        if filename.find('.pdf') == -1:
            # print(file_path)
            os.remove(file_path)
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    if path.exists(r'raksha/raksha' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'raksha/raksha' + str(sys.argv[6]) + '.xlsx')
    import openpyxl

    po = []
    wbkName = 'raksha/raksha' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    for t in range(0, len(onlyfiles)):
        try:
            sh1 = ['Sr No.', 'Claim No', 'Member Id', 'Dev/AgentCode', 'Insured/Employee', 'Claimant/Patient',
                   'Claim Amount', 'Pass Amount', 'Net Payable Amount', 'Payment', 'Diagnosis',
                   'Insurance Company Claim No.', 'Policy Number', 'Policy Period', 'Period of Hospitalization',
                   'Corporate Name', 'Deduction', 'TDS*', 'Neft-Ref/Cheque No.', 'Neft-Ref/Cheque Date', 'Hospital Name']
            sh2 = ['Sr No.', 'Claim ID', 'category', 'Billed Amt(Rs)', 'Deduction Amt(Rs)', 'Approved Amt(Rs)',
                   'Reason of Deduction (If any)']

            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 1).value = sh1[i]
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 1).value = sh2[i]
            tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all')
            tables.export('raksha/foo1.xls', f='excel')
            loc = ("raksha/foo1.xls")
            wb = xlrd.open_workbook(loc)
            sheet_1 = wb.sheet_by_index(0)
            sheet_1.cell_value(0, 0)
            s = []
            for k in range(1, len(sh1)):
                c = 0
                for j in [1, 3]:
                    for i in range(1, sheet_1.nrows):
                        if (sheet_1.cell_value(i, j) == sh1[k]):
                            # print(sh1[k],j)
                            if (sheet_1.cell_value(i, j) == 'Period of Hospitalization'):
                                if (j == 1):
                                    temp = sheet_1.cell_value(i, j + 1)
                                    temp1 = temp.find('DOD')
                                    doa = temp[:temp1]
                                    dod = temp[temp1 + 4:]
                                if (j == 3):
                                    temp = sheet_1.cell_value(i, j + 1)
                                    temp1 = temp.find('DOD')
                                    doa = temp[4:temp1]
                                    dod = temp[temp1 + 4:]
                            elif (j == 1):
                                c = 1
                                s.append(sheet_1.cell_value(i, 2))
                            elif (j == 3):
                                c = 1
                                s.append(sheet_1.cell_value(i, 4))
                if c == 0:
                    s.append(' ')
            s.append(doa)
            s.append(dod)
            # print(s)
            s1.cell(row=t + 2, column=1).value = t + 1
            for i in range(0, len(s)):
                s1.cell(row=t + 2, column=i + 2).value = s[i]
            hg = []
            gh = []
            h = []
            g = []
            hj = []
            for j in range(1, tables.n):
                sheet_n = wb.sheet_by_index(j)
                sheet_n.cell_value(0, 0)
                for i in range(2, sheet_n.nrows):
                    hg.append(sheet_n.cell_value(i, 1))
                    gh.append(sheet_n.cell_value(i, 2))
                    h.append(sheet_n.cell_value(i, 3))
                    g.append(sheet_n.cell_value(i, 4))
                    hj.append(sheet_n.cell_value(i, 5))
            for i in range(0, len(hg)):
                wq += 1
                row_num = s2.max_row
                s2.cell(row=row_num + 1, column=1).value = wq
                s2.cell(row=row_num + 1, column=2).value = s[0]
                s2.cell(row=row_num + 1, column=3).value = hg[i]
                s2.cell(row=row_num + 1, column=4).value = gh[i]
                s2.cell(row=row_num + 1, column=5).value = h[i]
                s2.cell(row=row_num + 1, column=6).value = g[i]
                s2.cell(row=row_num + 1, column=7).value = hj[i]
            os.rename(os.getcwd() + '/raksha/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/raksha/attachments_' + str(sys.argv[6]) + '/' + s[0] + '.pdf')
        except Exception as e:
            s1.cell(row=t + 2, column=1).value = 'error'
            os.rename(os.getcwd() + '/raksha/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/raksha/attachments_' + str(sys.argv[6]) + '/' + s[0] + '.pdf')

    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'raksha'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    wbk.close
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
	log_exceptions()
