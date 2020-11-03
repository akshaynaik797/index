import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
import sys
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import pdftotext
import html2text

from decode_error import check_subject, read_from_delete
from make_log import log_exceptions

if path.exists(r'health_india/health_india' + str(sys.argv[6]) + '.xlsx'):
    os.remove(r'health_india/health_india' + str(sys.argv[6]) + '.xlsx')
import openpyxl
import subprocess
try:

    po = []
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])
    wq = 0
    qw = 0
    eu = []
    fg = []
    repeat = []
    from openpyxl.styles import Color, PatternFill, Font, Border

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')


    def read_email_from_gmail():
        b = 0
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(FROM "hitpainfo@healthindiatpa.com" SUBJECT "Claim Settlement for Patient" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None,
        #                          '(FROM "hitpainfo@healthindiatpa.com" SUBJECT "Claim Settlement for Patient" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id, "(RFC822)")
            ##################################################ak
            try:
                raw_email = data[0][1].decode('utf-8')
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                b += 1
                for part in email_message.walk():
                    # print(part.get_content_type())
                    if part.get_content_type() == "text/html":
                        # print('hi')
                        body = part.get_payload(decode=True)
                        file_name = "health_india/email.html"
                        output_file = open(file_name, 'w')
                        output_file.write("Body: %s" % (body.decode('utf-8')))
                        output_file.close()

                        pdfkit.from_file('health_india/email.html',
                                         'health_india/attachments_' + str(sys.argv[6]) + '/' + str(b) + '.pdf')
                # df.to_csv('attachments/'+'%s'%i+'myfile_%s.csv'% t)
            else:
                repeat.append(email_message['Subject'])
            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/health_india'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/health_india/attachments_' + str(sys.argv[6]) + '/'

    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    wbkName = 'health_india/health_india' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')

    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    sh1 = ['Sr No.', 'Claim Number', 'Insurance Company', 'Policy Number', 'Corporate/Retail', 'Type of Claim',
           'Employee Code', 'Employee Name', 'Patient Name', 'Hospital Name', 'Hospital City',
           'Date of Admission - Discharge', 'Ailment', 'UTR No', 'UTR Date', 'Claim Amount', 'Deduction Amount',
           'Discount Amount', 'Approved Amount', 'TDS Amount', 'NEFT/Paid Amount']
    sh2 = ['Sr No.', 'Claim ID', 'Bill No.', 'Bill Date', 'Bill Amt', 'Payable Amt', 'Disallowance amount',
           'Disallowance Reasons', 'Deduction Category']

    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    for t in range(0, len(onlyfiles)):
        try:
            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 1).value = sh1[i]
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 1).value = sh2[i]
            tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all')
            tables.export('health_india/foo1.xls', f='excel')
            loc = ("health_india/foo1.xls")
            wb = xlrd.open_workbook(loc)
            s = []
            d = []
            sheet_3 = wb.sheet_by_index(0)
            sheet_3.cell_value(0, 0)

            for i in range(1, sheet_3.nrows - 1):
                d.append(sheet_3.cell_value(i, 1))
                s.append(sheet_3.cell_value(i, 2))
            ccn = s[0]
            s = [sub.replace('\t', ' ') for sub in s]
            d = [sub.replace('\t', ' ') for sub in d]
            d = [sub.replace('\n', ' ') for sub in d]
            s = [sub.replace('Rs.', '') for sub in s]
            s = [sub.replace('/-', '') for sub in s]
            # print(s)
            s1.cell(row=t + 2, column=1).value = t + 1
            for i in range(0, len(d)):
                k = sh1.index(d[i])
                s1.cell(row=t + 2, column=k + 1).value = s[i]
            sheet_1 = wb.sheet_by_index(1)
            sheet_1.cell_value(0, 0)
            d = []
            s = []
            p = []
            r = []
            e = []
            ro = []
            po = []
            for i in range(2, sheet_1.nrows):
                d.append(sheet_1.cell_value(i, 1))
                s.append(sheet_1.cell_value(i, 2))
                p.append(sheet_1.cell_value(i, 3))
                r.append(sheet_1.cell_value(i, 4))
                e.append(sheet_1.cell_value(i, 5))
                ro.append(sheet_1.cell_value(i, 6))
                po.append(sheet_1.cell_value(i, 7))
            if (tables.n == 3):
                sheet_2 = wb.sheet_by_index(2)
                sheet_2.cell_value(0, 0)
                for i in range(1, sheet_2.nrows):
                    d.append(sheet_2.cell_value(i, 1))
                    s.append(sheet_2.cell_value(i, 2))
                    p.append(sheet_2.cell_value(i, 3))
                    r.append(sheet_2.cell_value(i, 4))
                    e.append(sheet_2.cell_value(i, 5))
                    ro.append(sheet_2.cell_value(i, 6))
                    po.append(sheet_2.cell_value(i, 7))
            # print(p)
            num_ko = s1.max_row
            for i in range(0, len(d)):
                if (d[i] != ''):
                    if po[i].find('Discount') != -1:

                        for ko in range(1, num_ko + 1):
                            # print(e[i],ccn,s1.cell(row=ko, column=2).value)
                            if (ccn == s1.cell(row=ko, column=2).value):
                                s1.cell(row=ko, column=18).value = e[i]
                            # print(e[i])
                    row_num = s2.max_row + 1
                    wq += 1
                    s2.cell(row=row_num, column=1).value = wq
                    s2.cell(row=row_num, column=2).value = ccn
                    s2.cell(row=row_num, column=3).value = d[i]
                    s2.cell(row=row_num, column=4).value = s[i]
                    s2.cell(row=row_num, column=5).value = p[i]
                    s2.cell(row=row_num, column=6).value = r[i]
                    s2.cell(row=row_num, column=7).value = e[i]
                    s2.cell(row=row_num, column=8).value = ro[i]
                    s2.cell(row=row_num, column=9).value = po[i]

            os.rename(os.getcwd() + '/health_india/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/health_india/attachments_' + str(sys.argv[6]) + '/' + ccn + '.pdf')
        except Exception as e:
            s1.cell(row=t + 2, column=1).value = 'error'
            os.rename(os.getcwd() + '/health_india/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/health_india/attachments_' + str(sys.argv[6]) + '/' + ccn + '.pdf')

    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'health india'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()